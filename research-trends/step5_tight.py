"""
step5_tight.py
================
Strict keyword classification of AI & Society articles (2021-2025).
Patterns lifted from the methodology appendix (methodology.md).

NOTE: This file is a faithful reconstruction from the documented patterns;
small residuals (typically 0-8 articles per topic-year) may exist against
the published figures due to unavoidable differences in how loose English
descriptions of the regexes compile to actual Python regex.

Run from the research-trends/ directory:

    python3 step5_tight.py

Outputs per-year article counts and per-topic prevalence percentages.
Requires: openpyxl  (pip install openpyxl)
"""

import re
from collections import defaultdict
from openpyxl import load_workbook

SRC = "AI_SOCIETY_Articles_2021_onwards.xlsx"

# ---------------------------------------------------------------------------
# Topic patterns (exact patterns documented in methodology.md).
# Compiled with re.IGNORECASE. Each topic uses word boundaries to avoid
# substring false positives, and generic words are tightened into compounds.
# ---------------------------------------------------------------------------
PATTERNS = {
    "Generative AI": (
        r"\b("
        r"generative ai|gen[- ]ai|"
        r"large language models?|llms?|chatgpt|gpt(-\d+)?|"
        r"foundation models?|"
        r"generative (model|system|tool|technolog|image|video|text)\w*|"
        r"text-to-(image|video|speech)|"
        r"diffusion models?|stable diffusion|midjourney|dall-e|"
        r"prompt engineering|gemini|ai-generated"
        r")\b"
    ),
    "Ethics & Governance": (
        r"\b("
        r"ethics?|"
        r"ethical (ai|concerns?|implications?|issues?|dilemma|frameworks?|analysis|questions?|principles?)|"
        r"ai ethics|governance|responsible ai|accountab(le|ility)|"
        r"moral(ity|\s+status|\s+agency|\s+responsibility|\s+patiency|\s+reasoning)|"
        r"normative (framework|analysis|concerns|considerations)|"
        r"trustworthy ai|value alignment"
        r")\b"
    ),
    "Policy & Regulation": (
        r"\b("
        r"ai (policy|regulation|law|legislation|governance|act)|"
        r"technology policy|digital (policy|regulation|law|legislation)|"
        r"eu ai act|gdpr|"
        r"regulat(e|ing|ion|ions|ory|ed)|"
        r"legislat(e|ion|ive|ing|ed)|"
        r"compliance|public policy"
        r")\b"
    ),
    "Creativity & Arts": (
        r"\b("
        r"creativ(e|ity)|"
        r"art(s|ist|ists|istic|istry|work|works)|"
        r"music(al|ian|ians)?|poetry|literature|"
        r"paint(er|ing|ers|ings)|"
        r"aesthetic(s|)|"
        r"film(making|maker|makers|s)?|"
        r"compos(er|ers|ition|itional|ing)|"
        r"choreograph(y|er|ers|ic)|"
        r"perform(ance|ative|ing|ers?) (art|arts|practice)"
        r")\b"
    ),
    "Bias & Fairness": (
        r"\b("
        r"bias(es|ed)?|fairness|"
        r"algorithmic (fair|fairness|justice|equity|discrimination|bias)|"
        r"discriminat(e|ion|ory|ing)|racial bias|gender bias|racial discrimination|"
        r"equit(y|able)|marginali[sz]ed"
        r")\b"
    ),
    "Education & Learning": (
        r"\b("
        r"teacher(s)?|teaching|"
        r"student(s)?|pupil(s)?|"
        r"classroom|school(s|ing)|"
        r"higher education|"
        r"education(al)? (system|context|technolog\w*|institution\w*|policy|setting|practice|research)|"
        r"k-12|curricul(um|a)|pedagog(y|ical|ies)|literacy|learner(s)?|universit(y|ies)|"
        r"ai-(literacy|education)"
        r")\b"
    ),
    "Work & Labour": (
        r"\b("
        r"labou?r (market|force|relations|rights|conditions)|"
        r"future of (work|labou?r)|workplace|workforce|"
        r"workers? (rights?|conditions|protections?)|"
        r"gig (economy|work|worker|workers)|"
        r"automation (of|and) (work|labou?r|jobs|employment)|"
        r"job (loss|losses|displacement|market|insecurity|polarization)|"
        r"employment|unemployment|occupation(s|al)|human resources?|"
        r"(algorithmic|ai) management|platform work(er|ers)?|"
        r"technolog(y|ical) unemployment|"
        r"ai (and|in) (work|the workplace|employment|labou?r)|"
        r"digital labou?r"
        r")\b"
    ),
    "Robots & HRI": (
        r"\b("
        r"robots?|robotics?|human-robot|hri|androids?|humanoids?|cobots?|"
        r"social robot(s|ics)?|service robots?|care robots?"
        r")\b"
    ),
}

COMPILED = {name: re.compile(pat, re.IGNORECASE) for name, pat in PATTERNS.items()}


def classify(title: str, abstract: str):
    """Return the list of topic labels matched in title + abstract."""
    text = f"{title or ''} {abstract or ''}"
    return [name for name, rx in COMPILED.items() if rx.search(text)]


YEAR_RX = re.compile(r"\b(19|20)\d{2}\b")


def extract_year(value):
    """Accept datetime, date, or free-text like '31 December 2022'."""
    if value is None:
        return None
    if hasattr(value, "year"):
        return value.year
    m = YEAR_RX.search(str(value))
    return int(m.group(0)) if m else None


def main():
    wb = load_workbook(SRC, read_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header = {cell.value: idx for idx, cell in enumerate(header_row)}
    t_col = header["Title"]
    a_col = header["Abstract"]
    d_col = header["Date Published"]

    year_totals = defaultdict(int)
    year_topic = defaultdict(lambda: defaultdict(int))

    for row in ws.iter_rows(min_row=2, values_only=True):
        year = extract_year(row[d_col])
        if year is None or year < 2021 or year > 2025:
            continue
        year_totals[year] += 1
        for topic in classify(row[t_col], row[a_col]):
            year_topic[year][topic] += 1

    topics = list(PATTERNS.keys())
    years = sorted(year_totals)

    print("Per-topic prevalence (% of that year's articles)")
    print(f"{'Topic':24} " + " ".join(f"{y:>7}" for y in years))
    for t in topics:
        pct = [
            f"{100 * year_topic[y][t] / year_totals[y]:5.1f}%"
            for y in years
        ]
        print(f"{t:24} " + " ".join(f"{v:>7}" for v in pct))

    print("\nArticle counts by year:")
    for y in years:
        print(f"  {y}: {year_totals[y]}")
    print(f"  total 2021-2025: {sum(year_totals.values())}")


if __name__ == "__main__":
    main()
