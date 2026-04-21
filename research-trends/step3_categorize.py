"""
step3_categorize.py
====================
Permissive (loose) baseline for AI & Society topic prevalence.

This is kept alongside step5_tight.py for comparison. It intentionally uses
bare high-frequency words (e.g. 'work', 'learning', 'design') that the final
strict run deliberately excludes. The loose version inflates topic shares —
most dramatically for Work & Labour, which appears in ~42% of 2025 articles
under this pass versus 4.4% under step5_tight.py — because the bare word
'work' matches incidental phrases like "this work argues" and "framework".

Run from the research-trends/ directory:

    python3 step3_categorize.py

Requires: openpyxl  (pip install openpyxl)
"""

import re
from collections import defaultdict
from openpyxl import load_workbook

SRC = "AI_SOCIETY_Articles_2021_onwards.xlsx"

LOOSE = {
    "Generative AI": (
        r"\b(generative|gpt|chatgpt|llm|llms|large language|"
        r"foundation model|diffusion|midjourney|dall-e|"
        r"prompt|gemini|ai-generated)\b"
    ),
    "Ethics & Governance": (
        r"\b(ethic\w*|moral\w*|responsib\w*|accountab\w*|"
        r"trustworth\w*|governance|norm\w*)\b"
    ),
    "Policy & Regulation": (
        r"\b(policy|policies|regulat\w*|legislat\w*|"
        r"law|laws|gdpr|compliance|act)\b"
    ),
    "Creativity & Arts": (
        r"\b(creativ\w*|art|arts|artist\w*|music\w*|paint\w*|"
        r"film\w*|aesthetic\w*|design\w*|author\w*|writing|"
        r"culture|poem|poetry|literature)\b"
    ),
    "Bias & Fairness": (
        r"\b(bias\w*|fair\w*|equit\w*|discriminat\w*|"
        r"marginal\w*|justice|inequalit\w*)\b"
    ),
    "Education & Learning": (
        r"\b(teach\w*|student\w*|school\w*|class\w*|curricul\w*|"
        r"pedagog\w*|learn\w*|universit\w*|education\w*)\b"
    ),
    "Work & Labour": (
        r"\b(work\w*|labou?r\w*|employ\w*|unemploy\w*|job\w*|"
        r"worker\w*|occupation\w*|workforce|workplace|gig)\b"
    ),
    "Robots & HRI": (
        r"\b(robot\w*|android\w*|humanoid\w*|cobot\w*|hri)\b"
    ),
}

COMPILED = {name: re.compile(pat, re.IGNORECASE) for name, pat in LOOSE.items()}


def classify(title: str, abstract: str):
    text = f"{title or ''} {abstract or ''}"
    return [name for name, rx in COMPILED.items() if rx.search(text)]


YEAR_RX = re.compile(r"\b(19|20)\d{2}\b")


def extract_year(value):
    """Accept datetime, date, or free-text like '31 December 2022'."""
    if value is None:
        return None
    if hasattr(value, "year"):
        return value.year
    m = YEAR_RX.search(str(value))
    return int(m.group(0)) if m else None


def main():
    wb = load_workbook(SRC, read_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header = {cell.value: idx for idx, cell in enumerate(header_row)}
    t_col = header["Title"]
    a_col = header["Abstract"]
    d_col = header["Date Published"]

    year_totals = defaultdict(int)
    year_topic = defaultdict(lambda: defaultdict(int))

    for row in ws.iter_rows(min_row=2, values_only=True):
        year = extract_year(row[d_col])
        if year is None or year < 2021 or year > 2025:
            continue
        year_totals[year] += 1
        for topic in classify(row[t_col], row[a_col]):
            year_topic[year][topic] += 1

    topics = list(LOOSE.keys())
    years = sorted(year_totals)

    print("Per-topic prevalence (loose baseline) — % of that year's articles")
    print(f"{'Topic':24} " + " ".join(f"{y:>7}" for y in years))
    for t in topics:
        pct = [
            f"{100 * year_topic[y][t] / year_totals[y]:5.1f}%"
            for y in years
        ]
        print(f"{t:24} " + " ".join(f"{v:>7}" for v in pct))


if __name__ == "__main__":
    main()
